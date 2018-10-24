# -*- coding: utf-8 -*-
import pandas as pd
import os
import sys
# import xlwings as xw
from class_Data import Data
from openpyxl import load_workbook, Workbook
# from win32com import client
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
from logging_err import *


class Sheets(Data):
    Data_for_record = dict()
    # Шаблон записи по группам
    choose_param = list()

    # --------------------------------------------------------------------------------------------------------------


    # зачитка листа excel
    # --------------------------------------------------------------------------------------------------------------
    def read_sheet_by_excel(self, flag):
        try:
            self.DataFrame_Sheet = pd.read_excel(self.work_path_project,
                                                 sheet_name=self.workSheets_TC_TU,
                                                 skiprows=1,
                                                 usecols=self.work_cols)
            if flag:
                self.DataFrame_Sheet.loc[-1] = ''

        except Exception as e:
            exeption_print(e)
            return 0
    # --------------------------------------------------------------------------------------------------------------
    def choose_unic_part(self):

        # Сравниваем все варианты срезов, с вариантом который будет
        self.in_param = [x for x in self.in_param if x in self.DataFrame_Sheet.iloc[:, 2].tolist()]
        self.in_param.append(None)
        self.Dict_in_param = dict.fromkeys(self.in_param, pd.DataFrame())

    # --------------------------------------------------------------------------------------------------------------
    def index_col(self):
        self.DataFrame_Sheet = self.DataFrame_Sheet.set_index('Наименование команды ТУ')


    def format_sheet(self):

        self.choose_param = list(map(lambda x: set(self.in_param) & set(x), self.format_to_excel))

    def write_to_excel_win32com_pd(self, startrow, startcol, path, file_name, conv_file):
        Excel = client.Dispatch('Excel.Application')
        wb = Excel.Workbooks.Open(path + file_name)

        for Sheet in self.Data_for_record.keys():
            a = self.Data_for_record.get(Sheet)
            if not a.empty:
                sh = Excel.Worksheets(Sheet)
                sh.Range(sh.Cells(startrow, startcol),  # Cell to start the "paste"
                         sh.Cells(startrow + len(a.index) - 1,
                                  len(a.columns))
                         ).value = a.values.tolist()
                sh.Cells(4, 1).value = '{0}{1}'.format('Станция   ', conv_file)
        print(file_name)
        wb.Save()
        wb.Close()
        Excel.Quit()

    def write_to_excel(self, startrow, index, path, file_name, conv_file):

        mv = load_workbook(os.path.join(path, file_name))
        # print([lambda ws: ws.title, mv.worksheets])
        # print([mv.get_sheet_names()])
        # mv.template=True
        print(file_name)
        # for Sheet in self.Data_for_record.keys():
        #     sh=mv[Sheet]
        #     rows = dataframe_to_rows(self.Data_for_record.get(Sheet),index=index, header=False)
        #
        #     for r_idx, row in enumerate(rows, startrow+1):
        #         for c_idx, value in enumerate(row, 1):
        #             sh.cell(row=r_idx, column=c_idx, value=value)
        with pd.ExcelWriter(os.path.join(path, file_name), engine='openpyxl') as write_to_report:
            write_to_report.book = mv
            write_to_report.sheets = dict((ws.title, ws) for ws in mv.worksheets)
            for Sheet in self.Data_for_record.keys():
                self.Data_for_record.get(Sheet).to_excel(write_to_report,
                                                         sheet_name=Sheet,
                                                         startrow=startrow,
                                                         startcol=0,
                                                         header=False,
                                                         index=index)
                sh = mv[Sheet]
                sh.cell(row=4, column=1, value='{0}{1}'.format('Станция ', conv_file))
                # sh['A7'].border=Border(left=Side(border_style='thin'),
                #                          right=Side(style='thin'),
                #                          top=Side(style='thin'),
                #                          bottom=Side(style='thin'))
            write_to_report.save()
